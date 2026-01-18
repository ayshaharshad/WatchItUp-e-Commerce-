from django.utils import timezone
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, get_user_model
from django.db.models import Sum, Count, Q
from django.contrib.auth.decorators import user_passes_test, login_required
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.views.decorators.http import require_POST
from decimal import Decimal
from django.http import HttpResponse
from django.db.models.functions import TruncDate, TruncYear, TruncMonth
from datetime import datetime, timedelta


# For PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

# For Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO




# Models
from products.models import Category, Product, ProductImage, ProductVariant, ProductVariantImage, Order, OrderItem, OrderReturn, Coupon, CouponUsage, ProductOffer, CategoryOffer, ReferralCoupon, OrderCancellation
from users.models import Wallet, WalletTransaction
from .forms import (
    AdminLoginForm, CategoryForm, ProductForm, ProductVariantForm, 
    ProductVariantEditForm, BulkVariantForm, OrderStatusForm, OrderFilterForm, CouponForm, CouponFilterForm, ProductOfferForm, CategoryOfferForm, OfferFilterForm
)

User = get_user_model()



# ---------------- DECORATORS ----------------
def superuser_required(view_func):
    """
    Custom decorator that checks if user is authenticated AND is a superuser.
    Redirects to admin login if not authenticated or not a superuser.
    """
    decorated_view_func = login_required(login_url='admin_panel:admin_login')(view_func)
    return user_passes_test(
        lambda u: u.is_superuser,
        login_url='admin_panel:admin_login'
    )(decorated_view_func)

# ---------------- LOGIN ----------------
def admin_login(request):
    # Redirect to dashboard if already logged in as superuser
    if request.user.is_authenticated and request.user.is_superuser:
        return redirect("admin_panel:dashboard")
    
    if request.method == "POST":
        form = AdminLoginForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            if user.is_superuser:
                login(request, user)
                return redirect("admin_panel:dashboard")
            else:
                messages.error(request, "Only superusers can log in here.")
    else:
        form = AdminLoginForm()
    return render(request, "admin_panel/login.html", {"form": form})

@login_required(login_url="admin_panel:admin_login")
def admin_logout(request):
    logout(request)
    return redirect("admin_panel:admin_login")

# ---------------- DASHBOARD ----------------

@superuser_required
def dashboard(request):
    # Get filter parameters
    chart_filter = request.GET.get('chart_filter', 'monthly')
    
    # Basic statistics
    total_users = User.objects.filter(is_superuser=False).count()
    total_categories = Category.objects.filter(is_active=True).count()
    total_products = Product.objects.filter(is_active=True).count()
    blocked_users = User.objects.filter(is_superuser=False, is_active=False).count()
    
    # Calculate date range based on filter
    now = timezone.now()
    if chart_filter == 'yearly':
        # Last 12 months
        start_date = now - timedelta(days=365)
        date_trunc = TruncMonth
        date_format = '%b %Y'
    else:  # monthly (default)
        # Last 30 days
        start_date = now - timedelta(days=30)
        date_trunc = TruncMonth
        date_format = '%b %d'
    
    # Sales data for chart
    sales_data = Order.objects.filter(
        created_at__gte=start_date,
        status__in=['confirmed', 'processing', 'shipped', 'out_for_delivery', 'delivered']
    ).annotate(
        period=date_trunc('created_at')
    ).values('period').annotate(
        total_sales=Sum('total'),
        order_count=Count('id')
    ).order_by('period')
    
    # Format sales data for Chart.js
    chart_labels = []
    chart_sales = []
    chart_orders = []
    
    for data in sales_data:
        chart_labels.append(data['period'].strftime(date_format))
        chart_sales.append(float(data['total_sales'] or 0))
        chart_orders.append(data['order_count'])
    
    # Top 10 Bestselling Products
    top_products = OrderItem.objects.filter(
        order__status__in=['confirmed', 'processing', 'shipped', 'out_for_delivery', 'delivered'],
        status='active'
    ).values(
        'product__name',
        'product__id'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_revenue=Sum('item_total')
    ).order_by('-total_quantity')[:10]
    
    # Top 10 Bestselling Categories
    top_categories = OrderItem.objects.filter(
        order__status__in=['confirmed', 'processing', 'shipped', 'out_for_delivery', 'delivered'],
        status='active'
    ).values(
        'product__category__name',
        'product__category__id'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_revenue=Sum('item_total')
    ).order_by('-total_quantity')[:10]
    
    # Top 10 Bestselling Brands
    top_brands = OrderItem.objects.filter(
        order__status__in=['confirmed', 'processing', 'shipped', 'out_for_delivery', 'delivered'],
        status='active',
        product__brand__isnull=False
    ).values(
        'product__brand__name',
        'product__brand__id'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_revenue=Sum('item_total')
    ).order_by('-total_quantity')[:10]
    
    # Recent statistics
    total_revenue = Order.objects.filter(
        status__in=['confirmed', 'processing', 'shipped', 'out_for_delivery', 'delivered']
    ).aggregate(total=Sum('total'))['total'] or Decimal('0')
    
    total_orders = Order.objects.filter(
        status__in=['confirmed', 'processing', 'shipped', 'out_for_delivery', 'delivered']
    ).count()
    
    pending_orders = Order.objects.filter(status='pending').count()
    
    # Stock statistics
    low_stock_count = ProductVariant.objects.filter(
        is_active=True,
        stock_quantity__lte=5,
        stock_quantity__gt=0
    ).count()
    
    out_of_stock_count = ProductVariant.objects.filter(
        is_active=True,
        stock_quantity=0
    ).count()
    
    context = {
        # Basic stats
        'total_users': total_users,
        'total_categories': total_categories,
        'total_products': total_products,
        'blocked_users': blocked_users,
        
        # Sales stats
        'total_revenue': total_revenue,
        'total_orders': total_orders,
        'pending_orders': pending_orders,
        
        # Stock stats
        'low_stock_count': low_stock_count,
        'out_of_stock_count': out_of_stock_count,
        
        # Chart data
        'chart_filter': chart_filter,
        'chart_labels': chart_labels,
        'chart_sales': chart_sales,
        'chart_orders': chart_orders,
        
        # Top 10 lists
        'top_products': top_products,
        'top_categories': top_categories,
        'top_brands': top_brands,
    }
    
    return render(request, "admin_panel/dashboard.html", context)

# ---------------- USER MANAGEMENT ----------------
@superuser_required
def user_list(request):
    search = request.GET.get("q", "")
    users = User.objects.filter(is_superuser=False)
    if search:
        users = users.filter(Q(username__icontains=search) | Q(email__icontains=search))
    users = users.order_by("-id")
    paginator = Paginator(users, 10)
    page = request.GET.get("page")
    users = paginator.get_page(page)
    return render(request, "admin_panel/user_list.html", {"users": users, "search": search})

@superuser_required
@require_http_methods(["GET", "POST"])
def block_unblock_user(request, user_id):
    user = get_object_or_404(User, id=user_id)

    if user.is_superuser:
        messages.error(request, "Cannot block/unblock superuser accounts.")
        return redirect("admin_panel:user_list")

    if request.method == "POST":
        action = request.POST.get('action')

        if action == 'block':
            if user.is_active:
                user.is_active = False
                user.save()
                messages.success(request, f"User '{user.username}' has been blocked successfully.")
            else:
                messages.info(request, f"User '{user.username}' is already blocked.")

        elif action == 'unblock':
            if not user.is_active:
                user.is_active = True
                user.save()
                messages.success(request, f"User '{user.username}' has been unblocked successfully.")
            else:
                messages.info(request, f"User '{user.username}' is already active.")
        else:
            messages.error(request, "Invalid action specified.")
    else:
        user.is_active = not user.is_active
        user.save()
        status = "unblocked" if user.is_active else "blocked"
        messages.success(request, f"User '{user.username}' has been {status} successfully.")

    return redirect("admin_panel:user_list")



# ---------------- CATEGORY MANAGEMENT ----------------
@superuser_required
def category_list(request):
    search = request.GET.get("q", "")
    categories = Category.objects.filter(is_active=True)
    if search:
        categories = categories.filter(name__icontains=search)
    categories = categories.order_by("-id")
    paginator = Paginator(categories, 10)
    page = request.GET.get("page")
    categories = paginator.get_page(page)
    return render(request, "admin_panel/category_list.html", {"categories": categories, "search": search})

@superuser_required
def add_category(request):
    if request.method == "POST":
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Category added successfully.")
            return redirect("admin_panel:category_list")
    else:
        form = CategoryForm()
    return render(request, "admin_panel/add_category.html", {"form": form})

@superuser_required
def edit_category(request, pk):
    category = get_object_or_404(Category, pk=pk)
    form = CategoryForm(request.POST or None, instance=category)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Category updated successfully.")
        return redirect("admin_panel:category_list")
    return render(request, "admin_panel/edit_category.html", {"form": form, "category": category})

@superuser_required
def delete_category(request, pk):
    category = get_object_or_404(Category, pk=pk)
    category.is_active = False
    category.save()
    messages.success(request, "Category deleted successfully.")
    return redirect("admin_panel:category_list")

# ---------------- PRODUCT MANAGEMENT ----------------
@superuser_required
def product_list(request):
    search = request.GET.get("q", "")
    products = Product.objects.filter(is_active=True).select_related('category', 'brand').prefetch_related('variants')
    if search:
        products = products.filter(Q(name__icontains=search) | Q(description__icontains=search))
    products = products.order_by("-id")
    paginator = Paginator(products, 10)
    page = request.GET.get("page")
    products = paginator.get_page(page)
    return render(request, "admin_panel/product_list.html", {"products": products, "search": search})


@superuser_required
def add_product(request):
    if request.method == "POST":
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Save the product
                    product = form.save()
                    
                    # Handle multiple images (now potentially cropped)
                    uploaded_images = request.FILES.getlist("images")
                    if uploaded_images:
                        for img in uploaded_images:
                            # Optional: Validate image size/type if needed after cropping
                            if img.size > 5 * 1024 * 1024:  # Example: Limit to 5MB
                                raise ValueError("Image too large after cropping.")
                            ProductImage.objects.create(product=product, image=img)
                    
                    messages.success(
                        request, 
                        f"Product '{product.name}' added successfully! You can now add variants with specific colors, sizes, and prices."
                    )
                    return redirect("admin_panel:add_product_variant", product_id=product.id)
            
            except Exception as e:
                messages.error(request, f"Error saving product: {str(e)}")
                print(f"Product creation error: {e}")  # For debugging
        else:
            # (Existing error handling remains)
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    if field == '__all__':
                        error_messages.append(error)
                    else:
                        error_messages.append(f"{field.title()}: {error}")
            
            if error_messages:
                messages.error(request, "Please correct the following errors: " + "; ".join(error_messages))
            else:
                messages.error(request, "Please correct the errors below.")
    else:
        form = ProductForm()
    
    return render(request, "admin_panel/add_product.html", {"form": form})


@superuser_required
def edit_product(request, pk):
    """Edit product view with optional image upload (0-3 images)"""
    product = get_object_or_404(Product, pk=pk)

    if request.method == "POST":
        try:
            # Update base product information
            product.name = request.POST.get('name')
            product.category_id = request.POST.get('category')
            product.brand_id = request.POST.get('brand') or None
            product.base_price = request.POST.get('base_price')
            product.stock = request.POST.get('stock', 0)
            product.description = request.POST.get('description', '')

            product.save()

            # Handle main product images (0-3 images allowed)
            uploaded_images = request.FILES.getlist('images')
            
            # Validate image count
            if len(uploaded_images) > 3:
                messages.error(request, "You can upload a maximum of 3 images per product.")
                return redirect("admin_panel:edit_product", pk=pk)
            
            # Only replace images if new ones are uploaded
            if uploaded_images:
                # Clear existing ProductImage entries
                product.images.all().delete()
                
                # Create new ProductImage entries
                for idx, img in enumerate(uploaded_images):
                    is_primary = (idx == 0)  # First image is primary
                    ProductImage.objects.create(
                        product=product, 
                        image=img,
                        is_primary=is_primary
                    )

            messages.success(request, "Product updated successfully!")
            return redirect("admin_panel:product_list")

        except Exception as e:
            messages.error(request, f"Error: {str(e)}")

    form = ProductForm(instance=product)
    return render(request, "admin_panel/edit_product.html", {
        "form": form,
        "product": product
    })

@superuser_required
def delete_product(request, pk):
    product = get_object_or_404(Product, pk=pk)
    product.is_active = False
    product.save()
    messages.success(request, "Product deleted successfully.")
    return redirect("admin_panel:product_list")

# ---------------- PRODUCT VARIANT MANAGEMENT ----------------
@superuser_required
def variant_list(request):
    search = request.GET.get("q", "")
    product_filter = request.GET.get("product", "")
    color_filter = request.GET.get("color", "")
    
    variants = ProductVariant.objects.filter(is_active=True).select_related('product')
    
    if search:
        variants = variants.filter(
            Q(product__name__icontains=search) | 
            Q(sku__icontains=search) |
            Q(color__icontains=search)
        )
    
    if product_filter:
        variants = variants.filter(product_id=product_filter)
        
    if color_filter:
        variants = variants.filter(color=color_filter)
    
    variants = variants.order_by("-id")
    
    paginator = Paginator(variants, 15)
    page = request.GET.get("page")
    variants = paginator.get_page(page)
    
    products = Product.objects.filter(is_active=True).order_by('name')
    colors = ProductVariant.COLOR_CHOICES
    
    context = {
        'variants': variants,
        'search': search,
        'products': products,
        'colors': colors,
        'product_filter': product_filter,
        'color_filter': color_filter,
    }
    
    return render(request, "admin_panel/variant_list.html", context)

@superuser_required
def add_product_variant(request, product_id=None):
    """Fixed add_product_variant to handle variant images properly"""
    product = None
    if product_id:
        product = get_object_or_404(Product, pk=product_id, is_active=True)

    if request.method == "POST":
        form = ProductVariantForm(request.POST)
        if form.is_valid():
            variant = form.save()
            
            # Handle variant images (separate from product images)
            if "images" in request.FILES:
                for i, img in enumerate(request.FILES.getlist("images")):
                    ProductVariantImage.objects.create(
                        variant=variant, 
                        image=img,
                        is_primary=(i == 0)  # First image is primary
                    )
            messages.success(request, f"Variant '{variant}' added successfully.")
            return redirect("admin_panel:variant_list")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        initial_data = {}
        if product:
            initial_data['product'] = product
        form = ProductVariantForm(initial=initial_data)

    context = {
        'form': form,
        'product': product,
    }
    return render(request, "admin_panel/add_variant.html", context)


@superuser_required
def edit_product_variant(request, pk):
    """Fixed edit_product_variant view to handle variant images separately"""
    variant = get_object_or_404(ProductVariant, pk=pk)

    if request.method == "POST":
        form = ProductVariantEditForm(request.POST, instance=variant)
        if form.is_valid():
            form.save()
            
            # Handle variant images separately - only if new images are uploaded
            uploaded_images = request.FILES.getlist('images')
            if uploaded_images:
                # Clear only THIS variant's images, not the main product images
                variant.images.all().delete()
                for i, img in enumerate(uploaded_images):
                    ProductVariantImage.objects.create(
                        variant=variant,
                        image=img,
                        is_primary=(i == 0)  # First image is primary
                    )
            messages.success(request, "Variant updated successfully!")
            return redirect("admin_panel:variant_list")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = ProductVariantEditForm(instance=variant)

    context = {
        'form': form,
        'variant': variant,
    }
    return render(request, "admin_panel/edit_variant.html", context)

@superuser_required
def delete_product_variant(request, pk):
    variant = get_object_or_404(ProductVariant, pk=pk)
    variant.is_active = False
    variant.save()
    messages.success(request, f"Variant '{variant}' deleted successfully.")
    return redirect("admin_panel:variant_list")

@superuser_required
def bulk_create_variants(request):
    if request.method == "POST":
        form = BulkVariantForm(request.POST)
        if form.is_valid():
            product = form.cleaned_data['product']
            colors = form.cleaned_data['colors']
            base_price = form.cleaned_data['base_price']
            stock_quantity = form.cleaned_data['stock_quantity']
            
            created_count = 0
            with transaction.atomic():
                for color in colors:
                    ProductVariant.objects.create(
                        product=product,
                        color=color,
                        price=base_price,
                        stock_quantity=stock_quantity
                    )
                    created_count += 1
            
            messages.success(request, f"Successfully created {created_count} variants for '{product.name}'.")
            return redirect("admin_panel:variant_list")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = BulkVariantForm()
    
    return render(request, "admin_panel/bulk_create_variants.html", {"form": form})

@superuser_required
def variant_stock_update(request):
    if request.method == "POST":
        updates = request.POST.getlist('stock_updates')
        updated_count = 0
        
        for update in updates:
            if update.strip():
                try:
                    variant_id, new_stock = update.split(':')
                    variant = ProductVariant.objects.get(id=int(variant_id), is_active=True)
                    variant.stock_quantity = int(new_stock)
                    variant.save()
                    updated_count += 1
                except (ValueError, ProductVariant.DoesNotExist):
                    continue
        
        messages.success(request, f"Updated stock for {updated_count} variants.")
        return redirect("admin_panel:variant_list")
    
    low_stock_variants = ProductVariant.objects.filter(
        is_active=True,
        stock_quantity__lte=5
    ).select_related('product').order_by('stock_quantity')
    
    return render(request, "admin_panel/variant_stock_update.html", {
        'low_stock_variants': low_stock_variants
    })

@superuser_required  
def product_variant_detail(request, product_id):
    product = get_object_or_404(Product, pk=product_id, is_active=True)
    variants = product.variants.filter(is_active=True).prefetch_related('images').order_by('color')
    
    context = {
        'product': product,
        'variants': variants,
    }
    return render(request, "admin_panel/product_variant_detail.html", context)


# -------------------- ORDER MANAGEMENT --------------------


@superuser_required
def order_list(request):
    """✅ FIXED: List all orders with correct calculations"""
    # Get filter parameters
    search = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '')
    payment_filter = request.GET.get('payment_method', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    sort_by = request.GET.get('sort', '-created_at')
    
    # Base queryset
    orders = Order.objects.select_related('user').prefetch_related(
        'items',
        'items__product',
        'items__variant'
    ).all()
    
    # Apply search
    if search:
        orders = orders.filter(
            Q(order_id__icontains=search) |
            Q(user__username__icontains=search) |
            Q(user__email__icontains=search) |
            Q(shipping_full_name__icontains=search) |
            Q(shipping_phone__icontains=search)
        )
    
    # Apply filters
    if status_filter:
        orders = orders.filter(status=status_filter)
    
    if payment_filter:
        orders = orders.filter(payment_method=payment_filter)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            orders = orders.filter(created_at__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            orders = orders.filter(created_at__lte=date_to_obj)
        except ValueError:
            pass
    
    # Apply sorting
    valid_sort_fields = [
        'created_at', '-created_at',
        'total', '-total',
        'status', '-status',
        'order_id', '-order_id'
    ]
    if sort_by in valid_sort_fields:
        orders = orders.order_by(sort_by)
    else:
        orders = orders.order_by('-created_at')
    
    # ✅ FIX: Calculate statistics from ACTIVE items only
    total_orders = orders.count()
    
    # Calculate revenue from active items only
    total_revenue = Decimal('0')
    for order in orders:
        total_revenue += order.active_total
    
    # Pagination
    paginator = Paginator(orders, 20)
    page = request.GET.get('page')
    orders_page = paginator.get_page(page)
    
    # ✅ FIX: Add cached counts as regular attributes (Django templates don't allow underscore)
    for order in orders_page:
        # Cache these values to avoid multiple DB queries
        order.cached_active_count = order.items.filter(status='active').count()
        order.cached_cancelled_count = order.items.filter(status='cancelled').count()
        order.cached_returned_count = order.items.filter(status='returned').count()
        
        # Add display fields
        order.cached_display_total = order.active_total
        order.cached_refund_total = order.refunded_amount
    
    # Filter form
    filter_form = OrderFilterForm(request.GET)
    
    context = {
        'orders': orders_page,
        'filter_form': filter_form,
        'search': search,
        'status_filter': status_filter,
        'payment_filter': payment_filter,
        'date_from': date_from,
        'date_to': date_to,
        'sort_by': sort_by,
        'total_orders': total_orders,
        'total_revenue': total_revenue,  # ✅ Now correct - active items only
    }
    
    return render(request, 'admin_panel/order_list.html', context)



@superuser_required
def order_detail(request, order_id):
    """✅ FIXED: Enhanced order detail with proper item grouping"""
    order = get_object_or_404(
        Order.objects.select_related('user').prefetch_related(
            'items__product',
            'items__variant',
            'cancellations',
            'returns'
        ),
        order_id=order_id
    )
    
    # ✅ FIX: Use model properties to group items by status
    # These are already defined in the Order model as properties
    active_items = order.items.filter(status='active')
    cancelled_items = order.items.filter(status='cancelled')
    returned_items = order.items.filter(status='returned')
    
    # Get cancellation history
    cancellations = order.cancellations.all().order_by('-cancelled_at')
    
    # Get return requests
    return_requests = order.returns.all().order_by('-requested_at')
    
    # ✅ FIX: Calculate correct refund totals
    total_refunds = Decimal('0')
    for cancellation in cancellations:
        if cancellation.refund_status == 'processed':
            total_refunds += cancellation.refund_amount
    
    for return_req in return_requests.filter(refund_status='processed'):
        total_refunds += return_req.refund_amount
    
    context = {
        'order': order,
        
        # ✅ Item groupings (already filtered)
        'active_items': active_items,
        'cancelled_items': cancelled_items,
        'returned_items': returned_items,
        
        # Cancellation/return history
        'cancellations': cancellations,
        'return_requests': return_requests,
        
        # ✅ Statistics
        'active_items_count': active_items.count(),
        'cancelled_items_count': cancelled_items.count(),
        'returned_items_count': returned_items.count(),
        'total_refunds': total_refunds,
        
        # ✅ Note: active_subtotal, active_tax, active_total are accessed 
        # directly in template via order.active_subtotal etc (model properties)
    }
    
    return render(request, 'admin_panel/order_detail.html', context)



@superuser_required
def order_detail_enhanced(request, order_id):
    """
    ✅ ENHANCED: Order detail view with item-level cancellation/return tracking
    This replaces your existing order_detail view
    """
    order = get_object_or_404(
        Order.objects.select_related('user').prefetch_related(
            'items__product',
            'items__variant',
            'cancellations',
            'returns'
        ),
        order_id=order_id
    )
    
    # ✅ Group items by status
    active_items = order.items.filter(status='active')
    cancelled_items = order.items.filter(status='cancelled')
    returned_items = order.items.filter(status='returned')
    
    # ✅ Get cancellation history (both full order and item-level)
    cancellations = order.cancellations.all().order_by('-cancelled_at')
    full_order_cancellations = cancellations.filter(cancellation_type='full_order')
    item_cancellations = cancellations.filter(cancellation_type='single_item')
    
    # ✅ Get return requests
    return_requests = order.returns.all().order_by('-requested_at')
    
    # ✅ Calculate refund totals
    total_refunds = Decimal('0')
    for cancellation in cancellations:
        if cancellation.refund_status == 'processed':
            total_refunds += cancellation.refund_amount
    
    for return_req in return_requests.filter(refund_status='processed'):
        total_refunds += return_req.refund_amount
    
    context = {
        'order': order,
        
        # Item groupings
        'active_items': active_items,
        'cancelled_items': cancelled_items,
        'returned_items': returned_items,
        
        # Cancellation/return history
        'cancellations': cancellations,
        'full_order_cancellations': full_order_cancellations,
        'item_cancellations': item_cancellations,
        'return_requests': return_requests,
        
        # Statistics
        'active_items_count': active_items.count(),
        'cancelled_items_count': cancelled_items.count(),
        'returned_items_count': returned_items.count(),
        'total_refunds': total_refunds,
    }
    
    return render(request, 'admin_panel/order_detail_enhanced.html', context)


# ==================== ITEM-LEVEL CANCELLATION ====================




@superuser_required
def cancel_order_items_admin(request, order_id):
    """✅ Admin can cancel specific items from an order"""
    order = get_object_or_404(Order, order_id=order_id)
    
    if request.method == 'POST':
        item_ids = request.POST.getlist('item_ids[]')
        reason = request.POST.get('reason', '').strip()
        
        if not item_ids:
            messages.error(request, 'Please select at least one item to cancel.')
            return redirect('admin_panel:order_detail', order_id=order.order_id)
        
        if not reason:
            messages.error(request, 'Please provide a cancellation reason.')
            return redirect('admin_panel:order_detail', order_id=order.order_id)
        
        try:
            with transaction.atomic():
                items_to_cancel = OrderItem.objects.filter(
                    id__in=item_ids,
                    order=order,
                    status='active'
                )
                
                if not items_to_cancel.exists():
                    messages.error(request, 'No valid items found to cancel.')
                    return redirect('admin_panel:order_detail', order_id=order.order_id)
                
                total_refund = Decimal('0')
                cancelled_count = 0
                
                for item in items_to_cancel:
                    # Restore stock
                    if item.variant:
                        item.variant.stock_quantity += item.quantity
                        item.variant.save()
                    else:
                        item.product.stock += item.quantity
                        item.product.save()
                    
                    # ✅ FIX: Calculate refund (only if payment was completed)
                    refund_amount = Decimal('0')
                    if order.payment_status == 'completed':
                        refund_amount = item.item_total
                        total_refund += refund_amount
                    
                    # Update item status
                    item.status = 'cancelled'
                    item.save()
                    
                    # Create cancellation record
                    OrderCancellation.objects.create(
                        order=order,
                        order_item=item,
                        cancellation_type='single_item',
                        reason=f"Admin cancellation: {reason}",
                        refund_amount=refund_amount,
                        refund_status='processed' if refund_amount > 0 else 'not_applicable',
                        cancelled_by=request.user,
                        processed_at=timezone.now() if refund_amount > 0 else None
                    )
                    
                    cancelled_count += 1
                
                # ✅ FIX: Process refund to wallet
                if total_refund > 0:
                    wallet = Wallet.objects.get_or_create(user=order.user)[0]
                    wallet.add_money(
                        amount=total_refund,
                        transaction_type='credit_refund_cancel',
                        description=f'Refund for {cancelled_count} cancelled item(s) from order {order.order_id}',
                        reference_id=order.order_id
                    )
                
                # ✅ FIX: Check if ALL items are cancelled
                if not order.items.filter(status='active').exists():
                    order.status = 'cancelled'
                    order.cancelled_at = timezone.now()
                    if order.payment_status == 'completed':
                        order.payment_status = 'refunded'
                    order.save()
                
                # Add note to order
                timestamp = timezone.now().strftime('%Y-%m-%d %H:%M')
                note = f"[{timestamp}] Admin cancelled {cancelled_count} item(s) by {request.user.username}: {reason}"
                if total_refund > 0:
                    note += f" | Refund: ₹{total_refund}"
                
                if order.notes:
                    order.notes = f"{order.notes}\n{note}"
                else:
                    order.notes = note
                order.save()
                
                messages.success(
                    request,
                    f'Successfully cancelled {cancelled_count} item(s). '
                    f'{"₹" + str(total_refund) + " refunded to customer wallet." if total_refund > 0 else ""}'
                )
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            messages.error(request, f'Error cancelling items: {str(e)}')
        
        return redirect('admin_panel:order_detail', order_id=order.order_id)
    
    # GET request - show cancellation form
    active_items = order.items.filter(status='active')
    
    context = {
        'order': order,
        'active_items': active_items,
    }
    
    return render(request, 'admin_panel/cancel_order_items.html', context)


# ==================== ITEM-LEVEL RETURN PROCESSING ====================

@superuser_required
def process_item_return_admin(request, return_id):
    """
    ✅ NEW: Admin processes individual item returns
    """
    return_request = get_object_or_404(OrderReturn, pk=return_id)
    order = return_request.order
    
    if request.method == 'POST':
        action = request.POST.get('action')  # 'approve' or 'reject'
        admin_notes = request.POST.get('admin_notes', '').strip()
        
        try:
            with transaction.atomic():
                if action == 'approve':
                    # Approve return
                    return_request.status = 'approved'
                    return_request.reviewed_by = request.user
                    return_request.reviewed_at = timezone.now()
                    return_request.approved_at = timezone.now()
                    return_request.admin_notes = admin_notes
                    return_request.save()
                    
                    # Update order status
                    order.status = 'return_approved'
                    order.save()
                    
                    messages.success(
                        request,
                        f'Return request approved. Waiting for items to be returned. '
                        f'Change order status to "Returned" to complete the refund.'
                    )
                    
                elif action == 'reject':
                    rejection_reason = request.POST.get('rejection_reason', '').strip()
                    
                    if not rejection_reason:
                        messages.error(request, 'Please provide a rejection reason.')
                        return redirect('admin_panel:return_request_detail', pk=return_id)
                    
                    # Reject return
                    return_request.status = 'rejected'
                    return_request.reviewed_by = request.user
                    return_request.reviewed_at = timezone.now()
                    return_request.rejected_at = timezone.now()
                    return_request.rejection_reason = rejection_reason
                    return_request.admin_notes = admin_notes
                    return_request.save()
                    
                    # Restore order to delivered
                    order.status = 'delivered'
                    order.save()
                    
                    messages.success(request, 'Return request rejected successfully.')
                
                else:
                    messages.error(request, 'Invalid action.')
                    
        except Exception as e:
            import traceback
            traceback.print_exc()
            messages.error(request, f'Error processing return: {str(e)}')
        
        return redirect('admin_panel:return_request_detail', pk=return_id)
    
    # GET request - redirect to return detail page
    return redirect('admin_panel:return_request_detail', pk=return_id)


# ==================== CANCELLATION HISTORY ====================

@superuser_required
def cancellation_history(request):
    """
    ✅ NEW: View all order cancellations (both full orders and individual items)
    """
    search = request.GET.get('search', '').strip()
    cancellation_type_filter = request.GET.get('type', '')
    refund_status_filter = request.GET.get('refund_status', '')
    
    # Base queryset
    cancellations = OrderCancellation.objects.select_related(
        'order',
        'order__user',
        'order_item',
        'order_item__product',
        'cancelled_by'
    ).all().order_by('-cancelled_at')
    
    # Apply search
    if search:
        cancellations = cancellations.filter(
            Q(order__order_id__icontains=search) |
            Q(order__user__username__icontains=search) |
            Q(order__user__email__icontains=search) |
            Q(reason__icontains=search)
        )
    
    # Apply filters
    if cancellation_type_filter:
        cancellations = cancellations.filter(cancellation_type=cancellation_type_filter)
    
    if refund_status_filter:
        cancellations = cancellations.filter(refund_status=refund_status_filter)
    
    # Pagination
    paginator = Paginator(cancellations, 20)
    page = request.GET.get('page')
    cancellations_page = paginator.get_page(page)
    
    # Statistics
    total_cancellations = cancellations.count()
    full_order_cancellations = cancellations.filter(cancellation_type='full_order').count()
    item_cancellations = cancellations.filter(cancellation_type='single_item').count()
    total_refunded = cancellations.filter(
        refund_status='processed'
    ).aggregate(total=Sum('refund_amount'))['total'] or Decimal('0')
    
    context = {
        'cancellations': cancellations_page,
        'search': search,
        'cancellation_type_filter': cancellation_type_filter,
        'refund_status_filter': refund_status_filter,
        'total_cancellations': total_cancellations,
        'full_order_cancellations': full_order_cancellations,
        'item_cancellations': item_cancellations,
        'total_refunded': total_refunded,
    }
    
    return render(request, 'admin_panel/cancellation_history.html', context)


# ==================== CANCELLATION DETAIL ====================

@superuser_required
def cancellation_detail(request, cancellation_id):
    """
    ✅ NEW: View detailed information about a specific cancellation
    """
    cancellation = get_object_or_404(
        OrderCancellation.objects.select_related(
            'order',
            'order__user',
            'order_item',
            'order_item__product',
            'order_item__variant',
            'cancelled_by'
        ),
        pk=cancellation_id
    )
    
    context = {
        'cancellation': cancellation,
        'order': cancellation.order,
    }
    
    return render(request, 'admin_panel/cancellation_detail.html', context)


# ==================== AJAX: GET ORDER ITEMS ====================

@superuser_required
def get_order_items_for_cancellation(request, order_id):
    """
    ✅ NEW: AJAX endpoint to get active items for cancellation
    """
    order = get_object_or_404(Order, order_id=order_id)
    active_items = order.items.filter(status='active')
    
    items_data = []
    for item in active_items:
        items_data.append({
            'id': item.id,
            'product_name': item.product_name,
            'variant_name': item.variant_name or 'N/A',
            'quantity': item.quantity,
            'price': str(item.price),
            'item_total': str(item.item_total),
        })
    
    return JsonResponse({
        'success': True,
        'items': items_data,
        'can_cancel': order.can_cancel
    })


# ==================== STATISTICS DASHBOARD ====================

@superuser_required
def order_statistics_dashboard(request):
    """
    ✅ NEW: Comprehensive order statistics including cancellations/returns
    """
    from datetime import timedelta
    
    # Date range (default last 30 days)
    days = int(request.GET.get('days', 30))
    start_date = timezone.now() - timedelta(days=days)
    
    # Order statistics
    total_orders = Order.objects.count()
    orders_in_period = Order.objects.filter(created_at__gte=start_date).count()
    
    # Status breakdown
    status_breakdown = Order.objects.values('status').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Cancellation statistics
    total_cancellations = OrderCancellation.objects.count()
    recent_cancellations = OrderCancellation.objects.filter(
        cancelled_at__gte=start_date
    ).count()
    
    full_order_cancellations = OrderCancellation.objects.filter(
        cancellation_type='full_order'
    ).count()
    
    item_cancellations = OrderCancellation.objects.filter(
        cancellation_type='single_item'
    ).count()
    
    # Return statistics
    total_returns = OrderReturn.objects.count()
    pending_returns = OrderReturn.objects.filter(status='pending').count()
    approved_returns = OrderReturn.objects.filter(status='approved').count()
    completed_returns = OrderReturn.objects.filter(status='completed').count()
    rejected_returns = OrderReturn.objects.filter(status='rejected').count()
    
    # Financial statistics
    total_refunded = OrderCancellation.objects.filter(
        refund_status='processed'
    ).aggregate(total=Sum('refund_amount'))['total'] or Decimal('0')
    
    return_refunds = OrderReturn.objects.filter(
        refund_status='processed'
    ).aggregate(total=Sum('refund_amount'))['total'] or Decimal('0')
    
    total_refunds = total_refunded + return_refunds
    
    # Top cancellation reasons
    cancellation_reasons = OrderCancellation.objects.values('reason').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # Top return reasons
    return_reasons = OrderReturn.objects.values('reason').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    context = {
        'days': days,
        'start_date': start_date,
        
        # Order stats
        'total_orders': total_orders,
        'orders_in_period': orders_in_period,
        'status_breakdown': status_breakdown,
        
        # Cancellation stats
        'total_cancellations': total_cancellations,
        'recent_cancellations': recent_cancellations,
        'full_order_cancellations': full_order_cancellations,
        'item_cancellations': item_cancellations,
        
        # Return stats
        'total_returns': total_returns,
        'pending_returns': pending_returns,
        'approved_returns': approved_returns,
        'completed_returns': completed_returns,
        'rejected_returns': rejected_returns,
        
        # Financial stats
        'total_refunded': total_refunded,
        'return_refunds': return_refunds,
        'total_refunds': total_refunds,
        
        # Top reasons
        'cancellation_reasons': cancellation_reasons,
        'return_reasons': return_reasons,
    }
    
    return render(request, 'admin_panel/order_statistics.html', context)


@superuser_required
def update_order_status(request, order_id):
    """✅ FIXED: Update order status with proper refund handling"""
    order = get_object_or_404(Order, order_id=order_id)
    
    if request.method == 'POST':
        form = OrderStatusForm(request.POST, current_status=order.status)
        if form.is_valid():
            new_status = form.cleaned_data['status']
            notes = form.cleaned_data.get('notes', '')
            
            # Update order status
            old_status = order.status
            order.status = new_status
            
            # 1. DELIVERED STATUS - Set delivered timestamp
            if new_status == 'delivered' and not order.delivered_at:
                order.delivered_at = timezone.now()
            
            # 2. CANCELLED STATUS - Restore stock + refund
            elif new_status == 'cancelled' and old_status != 'cancelled':
                # ✅ FIX: Only restore stock for ACTIVE items
                for item in order.items.filter(status='active'):
                    if item.variant:
                        item.variant.stock_quantity += item.quantity
                        item.variant.save()
                    else:
                        item.product.stock += item.quantity
                        item.product.save()
                    
                    item.status = 'cancelled'
                    item.save()
                
                order.cancelled_at = timezone.now()
                
                # ✅ FIX: Refund only ACTIVE items amount
                if order.payment_status == 'completed':
                    refund_amount = order.active_total  # Only active items
                    order.payment_status = 'refunded'
                    
                    # Refund to wallet
                    from users.models import Wallet
                    wallet = Wallet.objects.get_or_create(user=order.user)[0]
                    wallet.add_money(
                        amount=refund_amount,
                        transaction_type='credit_refund_cancellation',
                        description=f'Refund for cancelled order {order.order_id}',
                        reference_id=order.order_id
                    )
            
            # 3. RETURN_APPROVED STATUS
            elif new_status == 'return_approved' and old_status == 'return_requested':
                return_request = order.returns.filter(status='pending').first()
                if return_request:
                    return_request.status = 'approved'
                    return_request.approved_at = timezone.now()
                    return_request.reviewed_by = request.user
                    return_request.reviewed_at = timezone.now()
                    return_request.save()
            
            # 4. ✅ RETURNED STATUS - Complete return + refund
            elif new_status == 'returned' and old_status in ['return_approved', 'return_requested', 'delivered']:
                # ✅ FIX: Restore stock only for ACTIVE items
                for item in order.items.filter(status='active'):
                    if item.variant:
                        item.variant.stock_quantity += item.quantity
                        item.variant.save()
                    else:
                        item.product.stock += item.quantity
                        item.product.save()
                    
                    item.status = 'returned'
                    item.save()
                
                # ✅ FIX: Refund only ACTIVE items
                refund_amount = order.active_total
                order.payment_status = 'refunded'
                
                # Refund to wallet
                from users.models import Wallet
                wallet = Wallet.objects.get_or_create(user=order.user)[0]
                wallet.add_money(
                    amount=refund_amount,
                    transaction_type='credit_refund_return',
                    description=f'Refund for returned order {order.order_id}',
                    reference_id=order.order_id
                )
                
                # Update return request
                return_request = order.returns.filter(
                    status__in=['pending', 'approved']
                ).first()
                if return_request:
                    return_request.status = 'completed'
                    return_request.completed_at = timezone.now()
                    return_request.refund_status = 'processed'
                    return_request.refund_amount = refund_amount  # ✅ Set correct amount
                    return_request.reviewed_by = request.user
                    return_request.reviewed_at = timezone.now()
                    return_request.save()
            
            # 5. If order was cancelled and now being reactivated
            if old_status == 'cancelled' and new_status in ['confirmed', 'pending', 'processing']:
                for item in order.items.filter(status='cancelled'):
                    if item.variant:
                        if item.variant.stock_quantity >= item.quantity:
                            item.variant.stock_quantity -= item.quantity
                            item.variant.save()
                            item.status = 'active'
                            item.save()
                        else:
                            messages.error(
                                request, 
                                f"Insufficient stock for {item.product_name}"
                            )
                            return redirect('admin_panel:order_detail', order_id=order.order_id)
                    else:
                        if item.product.stock >= item.quantity:
                            item.product.stock -= item.quantity
                            item.product.save()
                            item.status = 'active'
                            item.save()
                        else:
                            messages.error(
                                request, 
                                f"Insufficient stock for {item.product_name}"
                            )
                            return redirect('admin_panel:order_detail', order_id=order.order_id)
            
            # Add notes if provided
            if notes:
                timestamp = timezone.now().strftime('%Y-%m-%d %H:%M')
                new_note = f"[{timestamp}] Status changed from {old_status} to {new_status} by {request.user.username}: {notes}"
                if order.notes:
                    order.notes = f"{order.notes}\n{new_note}"
                else:
                    order.notes = new_note
            
            order.save()
            
            messages.success(
                request, 
                f"Order {order.order_id} status updated to {order.get_status_display()}"
            )
            
            return redirect('admin_panel:order_detail', order_id=order.order_id)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = OrderStatusForm(current_status=order.status)
    
    context = {
        'order': order,
        'form': form,
    }
    
    return render(request, 'admin_panel/update_order_status.html', context)



@superuser_required
def cancel_order(request, order_id):
    """Cancel an order and restore stock"""
    order = get_object_or_404(Order, order_id=order_id)
    
    if not order.can_cancel:
        messages.error(request, "This order cannot be cancelled.")
        return redirect('admin_panel:order_detail', order_id=order.order_id)
    
    if request.method == 'POST':
        reason = request.POST.get('reason', 'Cancelled by admin')
        
        # Update order status
        order.status = 'cancelled'
        order.notes = f"{order.notes}\n[{timezone.now().strftime('%Y-%m-%d %H:%M')}] Cancelled: {reason}" if order.notes else f"Cancelled: {reason}"
        order.save()
        
        # Restore stock
        for item in order.items.all():
            if item.variant:
                item.variant.stock_quantity += item.quantity
                item.variant.save()
        
        # Create cancellation record
        from products.models import OrderCancellation
        OrderCancellation.objects.create(
            order=order,
            reason=reason,
            cancelled_by=request.user
        )
        
        messages.success(request, f"Order {order.order_id} has been cancelled and stock restored.")
        return redirect('admin_panel:order_detail', order_id=order.order_id)
    
    context = {
        'order': order,
    }
    
    return render(request, 'admin_panel/cancel_order.html', context)


@superuser_required
def order_inventory_report(request):
    """View inventory/stock management report"""
    # Get all variants with stock information
    variants = ProductVariant.objects.filter(
        is_active=True
    ).select_related('product').order_by('stock_quantity')
    
    # Low stock variants (stock <= 10)
    low_stock = variants.filter(stock_quantity__lte=10)
    
    # Out of stock variants
    out_of_stock = variants.filter(stock_quantity=0)
    
    # Get recent orders to show stock movement
    recent_orders = Order.objects.filter(
        created_at__gte=timezone.now() - timezone.timedelta(days=30)
    ).select_related('user').order_by('-created_at')[:10]
    
    # Calculate stock statistics
    total_variants = variants.count()
    low_stock_count = low_stock.count()
    out_of_stock_count = out_of_stock.count()
    total_stock_value = sum(v.price * v.stock_quantity for v in variants)
    
    context = {
        'low_stock': low_stock,
        'out_of_stock': out_of_stock,
        'recent_orders': recent_orders,
        'total_variants': total_variants,
        'low_stock_count': low_stock_count,
        'out_of_stock_count': out_of_stock_count,
        'total_stock_value': total_stock_value,
    }
    
    return render(request, 'admin_panel/order_inventory_report.html', context)


@superuser_required
def clear_order_filters(request):
    """Clear all order filters and redirect to order list"""
    return redirect('admin_panel:order_list')


# -------------------- COUPON MANAGEMENT --------------------

@superuser_required
def coupon_list(request):
    """List all coupons with search and filters"""
    search = request.GET.get('search', '').strip()
    discount_type_filter = request.GET.get('discount_type', '')
    status_filter = request.GET.get('status', '')
    
    # Base queryset
    coupons = Coupon.objects.all().order_by('-created_at')
    
    # Apply search
    if search:
        coupons = coupons.filter(code__icontains=search)
    
    # Apply discount type filter
    if discount_type_filter:
        coupons = coupons.filter(discount_type=discount_type_filter)
    
    # Apply status filter
    now = timezone.now()
    if status_filter == 'active':
        coupons = coupons.filter(
            is_active=True,
            valid_from__lte=now,
            valid_to__gte=now
        )
    elif status_filter == 'inactive':
        coupons = coupons.filter(is_active=False)
    elif status_filter == 'expired':
        coupons = coupons.filter(valid_to__lt=now)
    
    # Add usage statistics for each coupon
    for coupon in coupons:
        coupon.total_usage = CouponUsage.objects.filter(coupon=coupon).count()
        coupon.total_discount_given = CouponUsage.objects.filter(
            coupon=coupon
        ).aggregate(total=Sum('discount_amount'))['total'] or 0
    
    # Pagination
    paginator = Paginator(coupons, 15)
    page = request.GET.get('page')
    coupons_page = paginator.get_page(page)
    
    # Filter form
    filter_form = CouponFilterForm(request.GET)
    
    context = {
        'coupons': coupons_page,
        'filter_form': filter_form,
        'search': search,
        'discount_type_filter': discount_type_filter,
        'status_filter': status_filter,
    }
    
    return render(request, 'admin_panel/coupon_list.html', context)


@superuser_required
def create_coupon(request):
    """Create a new coupon"""
    if request.method == 'POST':
        form = CouponForm(request.POST)
        if form.is_valid():
            coupon = form.save()
            messages.success(
                request,
                f"Coupon '{coupon.code}' created successfully!"
            )
            return redirect('admin_panel:coupon_list')
        else:
            # Display form errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        # Set default values for new coupon
        initial_data = {
            'is_active': True,
            'usage_per_user': 1,
            'discount_type': 'percentage',
            'valid_from': timezone.now(),
            'valid_to': timezone.now() + timezone.timedelta(days=30),
        }
        form = CouponForm(initial=initial_data)
    
    context = {
        'form': form,
        'title': 'Create New Coupon',
        'button_text': 'Create Coupon',
    }
    
    return render(request, 'admin_panel/coupon_form.html', context)


@superuser_required
def edit_coupon(request, pk):
    """Edit existing coupon"""
    coupon = get_object_or_404(Coupon, pk=pk)
    
    if request.method == 'POST':
        form = CouponForm(request.POST, instance=coupon)
        if form.is_valid():
            coupon = form.save()
            messages.success(
                request,
                f"Coupon '{coupon.code}' updated successfully!"
            )
            return redirect('admin_panel:coupon_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = CouponForm(instance=coupon)
    
    context = {
        'form': form,
        'coupon': coupon,
        'title': f'Edit Coupon: {coupon.code}',
        'button_text': 'Update Coupon',
    }
    
    return render(request, 'admin_panel/coupon_form.html', context)


@superuser_required
def delete_coupon(request, pk):
    """Soft delete a coupon (deactivate)"""
    coupon = get_object_or_404(Coupon, pk=pk)
    
    if request.method == 'POST':
        # Check if coupon has been used
        usage_count = CouponUsage.objects.filter(coupon=coupon).count()
        
        if usage_count > 0:
            # Soft delete - just deactivate
            coupon.is_active = False
            coupon.save()
            messages.success(
                request,
                f"Coupon '{coupon.code}' has been deactivated (it was used {usage_count} time(s))."
            )
        else:
            # Hard delete if never used
            code = coupon.code
            coupon.delete()
            messages.success(
                request,
                f"Coupon '{code}' has been permanently deleted."
            )
        
        return redirect('admin_panel:coupon_list')
    
    # Get usage statistics
    usage_count = CouponUsage.objects.filter(coupon=coupon).count()
    total_discount = CouponUsage.objects.filter(
        coupon=coupon
    ).aggregate(total=Sum('discount_amount'))['total'] or 0
    
    context = {
        'coupon': coupon,
        'usage_count': usage_count,
        'total_discount': total_discount,
    }
    
    return render(request, 'admin_panel/coupon_delete_confirm.html', context)


@superuser_required
def coupon_detail(request, pk):
    """View coupon details and usage statistics"""
    coupon = get_object_or_404(Coupon, pk=pk)
    
    # Get usage history
    usages = CouponUsage.objects.filter(coupon=coupon).select_related(
        'user', 'order'
    ).order_by('-used_at')
    
    # Pagination for usage history
    paginator = Paginator(usages, 20)
    page = request.GET.get('page')
    usages_page = paginator.get_page(page)
    
    # Calculate statistics
    total_usage = usages.count()
    total_discount_given = usages.aggregate(
        total=Sum('discount_amount')
    )['total'] or 0
    unique_users = usages.values('user').distinct().count()
    
    context = {
        'coupon': coupon,
        'usages': usages_page,
        'total_usage': total_usage,
        'total_discount_given': total_discount_given,
        'unique_users': unique_users,
    }
    
    return render(request, 'admin_panel/coupon_detail.html', context)


@superuser_required
def toggle_coupon_status(request, pk):
    """Toggle coupon active status"""
    coupon = get_object_or_404(Coupon, pk=pk)
    
    if request.method == 'POST':
        coupon.is_active = not coupon.is_active
        coupon.save()
        
        status = "activated" if coupon.is_active else "deactivated"
        messages.success(request, f"Coupon '{coupon.code}' has been {status}.")
    
    return redirect('admin_panel:coupon_list')


# ==================== PRODUCT OFFER MANAGEMENT ====================

@superuser_required
def product_offer_list(request):
    """List all product offers with filters"""
    search = request.GET.get('search', '').strip()
    discount_type_filter = request.GET.get('discount_type', '')
    status_filter = request.GET.get('status', '')
    
    # Base queryset
    offers = ProductOffer.objects.select_related('product').all().order_by('-created_at')
    
    # Apply search
    if search:
        offers = offers.filter(
            Q(name__icontains=search) |
            Q(product__name__icontains=search)
        )
    
    # Apply discount type filter
    if discount_type_filter:
        offers = offers.filter(discount_type=discount_type_filter)
    
    # Apply status filter
    now = timezone.now()
    if status_filter == 'active':
        offers = offers.filter(
            is_active=True,
            start_date__lte=now,
            end_date__gte=now
        )
    elif status_filter == 'inactive':
        offers = offers.filter(is_active=False)
    elif status_filter == 'expired':
        offers = offers.filter(end_date__lt=now)
    elif status_filter == 'upcoming':
        offers = offers.filter(start_date__gt=now)
    
    # Pagination
    paginator = Paginator(offers, 15)
    page = request.GET.get('page')
    offers_page = paginator.get_page(page)
    
    # Filter form
    filter_form = OfferFilterForm(request.GET)
    
    context = {
        'offers': offers_page,
        'filter_form': filter_form,
        'search': search,
        'discount_type_filter': discount_type_filter,
        'status_filter': status_filter,
        'offer_type': 'product',
    }
    
    return render(request, 'admin_panel/offer_list.html', context)


@superuser_required
def create_product_offer(request):
    """Create new product offer"""
    if request.method == 'POST':
        form = ProductOfferForm(request.POST)
        if form.is_valid():
            offer = form.save()
            messages.success(
                request,
                f"Product offer '{offer.name}' created successfully!"
            )
            return redirect('admin_panel:product_offer_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        # Set default values
        initial_data = {
            'is_active': True,
            'discount_type': 'percentage',
            'start_date': timezone.now(),
            'end_date': timezone.now() + timezone.timedelta(days=30),
        }
        form = ProductOfferForm(initial=initial_data)
    
    context = {
        'form': form,
        'title': 'Create Product Offer',
        'button_text': 'Create Offer',
        'offer_type': 'product',
    }
    
    return render(request, 'admin_panel/offer_form.html', context)


@superuser_required
def edit_product_offer(request, pk):
    """Edit existing product offer"""
    offer = get_object_or_404(ProductOffer, pk=pk)
    
    if request.method == 'POST':
        form = ProductOfferForm(request.POST, instance=offer)
        if form.is_valid():
            offer = form.save()
            messages.success(
                request,
                f"Product offer '{offer.name}' updated successfully!"
            )
            return redirect('admin_panel:product_offer_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = ProductOfferForm(instance=offer)
    
    context = {
        'form': form,
        'offer': offer,
        'title': f'Edit Product Offer: {offer.name}',
        'button_text': 'Update Offer',
        'offer_type': 'product',
    }
    
    return render(request, 'admin_panel/offer_form.html', context)


@superuser_required
def delete_product_offer(request, pk):
    """Delete product offer"""
    offer = get_object_or_404(ProductOffer, pk=pk)
    
    if request.method == 'POST':
        offer_name = offer.name
        offer.delete()
        messages.success(request, f"Product offer '{offer_name}' deleted successfully!")
        return redirect('admin_panel:product_offer_list')
    
    context = {
        'offer': offer,
        'offer_type': 'product',
    }
    
    return render(request, 'admin_panel/offer_delete_confirm.html', context)


@superuser_required
def toggle_product_offer_status(request, pk):
    """Toggle product offer active status"""
    offer = get_object_or_404(ProductOffer, pk=pk)
    
    if request.method == 'POST':
        offer.is_active = not offer.is_active
        offer.save()
        
        status = "activated" if offer.is_active else "deactivated"
        messages.success(request, f"Product offer '{offer.name}' has been {status}.")
    
    return redirect('admin_panel:product_offer_list')


# ==================== CATEGORY OFFER MANAGEMENT ====================

@superuser_required
def category_offer_list(request):
    """List all category offers with filters"""
    search = request.GET.get('search', '').strip()
    discount_type_filter = request.GET.get('discount_type', '')
    status_filter = request.GET.get('status', '')
    
    # Base queryset
    offers = CategoryOffer.objects.select_related('category').all().order_by('-created_at')
    
    # Apply search
    if search:
        offers = offers.filter(
            Q(name__icontains=search) |
            Q(category__name__icontains=search)
        )
    
    # Apply discount type filter
    if discount_type_filter:
        offers = offers.filter(discount_type=discount_type_filter)
    
    # Apply status filter
    now = timezone.now()
    if status_filter == 'active':
        offers = offers.filter(
            is_active=True,
            start_date__lte=now,
            end_date__gte=now
        )
    elif status_filter == 'inactive':
        offers = offers.filter(is_active=False)
    elif status_filter == 'expired':
        offers = offers.filter(end_date__lt=now)
    elif status_filter == 'upcoming':
        offers = offers.filter(start_date__gt=now)
    
    # Pagination
    paginator = Paginator(offers, 15)
    page = request.GET.get('page')
    offers_page = paginator.get_page(page)
    
    # Filter form
    filter_form = OfferFilterForm(request.GET)
    
    context = {
        'offers': offers_page,
        'filter_form': filter_form,
        'search': search,
        'discount_type_filter': discount_type_filter,
        'status_filter': status_filter,
        'offer_type': 'category',
    }
    
    return render(request, 'admin_panel/offer_list.html', context)


@superuser_required
def create_category_offer(request):
    """Create new category offer"""
    if request.method == 'POST':
        form = CategoryOfferForm(request.POST)
        if form.is_valid():
            offer = form.save()
            messages.success(
                request,
                f"Category offer '{offer.name}' created successfully!"
            )
            return redirect('admin_panel:category_offer_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        # Set default values
        initial_data = {
            'is_active': True,
            'discount_type': 'percentage',
            'start_date': timezone.now(),
            'end_date': timezone.now() + timezone.timedelta(days=30),
        }
        form = CategoryOfferForm(initial=initial_data)
    
    context = {
        'form': form,
        'title': 'Create Category Offer',
        'button_text': 'Create Offer',
        'offer_type': 'category',
    }
    
    return render(request, 'admin_panel/offer_form.html', context)


@superuser_required
def edit_category_offer(request, pk):
    """Edit existing category offer"""
    offer = get_object_or_404(CategoryOffer, pk=pk)
    
    if request.method == 'POST':
        form = CategoryOfferForm(request.POST, instance=offer)
        if form.is_valid():
            offer = form.save()
            messages.success(
                request,
                f"Category offer '{offer.name}' updated successfully!"
            )
            return redirect('admin_panel:category_offer_list')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = CategoryOfferForm(instance=offer)
    
    context = {
        'form': form,
        'offer': offer,
        'title': f'Edit Category Offer: {offer.name}',
        'button_text': 'Update Offer',
        'offer_type': 'category',
    }
    
    return render(request, 'admin_panel/offer_form.html', context)


@superuser_required
def delete_category_offer(request, pk):
    """Delete category offer"""
    offer = get_object_or_404(CategoryOffer, pk=pk)
    
    if request.method == 'POST':
        offer_name = offer.name
        offer.delete()
        messages.success(request, f"Category offer '{offer_name}' deleted successfully!")
        return redirect('admin_panel:category_offer_list')
    
    context = {
        'offer': offer,
        'offer_type': 'category',
    }
    
    return render(request, 'admin_panel/offer_delete_confirm.html', context)


@superuser_required
def toggle_category_offer_status(request, pk):
    """Toggle category offer active status"""
    offer = get_object_or_404(CategoryOffer, pk=pk)
    
    if request.method == 'POST':
        offer.is_active = not offer.is_active
        offer.save()
        
        status = "activated" if offer.is_active else "deactivated"
        messages.success(request, f"Category offer '{offer.name}' has been {status}.")
    
    return redirect('admin_panel:category_offer_list')


# ==================== REFERRAL MANAGEMENT ====================

@superuser_required
def referral_list(request):
    """View all referral activity"""
    search = request.GET.get('search', '').strip()
    
    # Get referral coupons
    referrals = ReferralCoupon.objects.select_related(
        'referrer', 'referred_user', 'coupon'
    ).all().order_by('-created_at')
    
    # Apply search
    if search:
        referrals = referrals.filter(
            Q(referrer__username__icontains=search) |
            Q(referrer__email__icontains=search) |
            Q(referred_user__username__icontains=search) |
            Q(referred_user__email__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(referrals, 20)
    page = request.GET.get('page')
    referrals_page = paginator.get_page(page)
    
    # Statistics
    total_referrals = referrals.count()
    active_referrals = referrals.filter(is_used=False).count()
    used_referrals = referrals.filter(is_used=True).count()
    
    # Top referrers
    from django.db.models import Count
    top_referrers = User.objects.filter(
        referral_coupons_earned__isnull=False
    ).annotate(
    total_referrals_earned=Count('referral_coupons_earned')
    ).order_by('-total_referrals_earned')
    
    context = {
        'referrals': referrals_page,
        'search': search,
        'total_referrals': total_referrals,
        'active_referrals': active_referrals,
        'used_referrals': used_referrals,
        'top_referrers': top_referrers,
    }
    
    return render(request, 'admin_panel/referral_list.html', context)



# ==================== RETURN REQUEST MANAGEMENT ====================

@superuser_required
def return_request_list(request):
    """List all return requests with filters"""
    search = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '')
    
    # Base queryset
    return_requests = OrderReturn.objects.select_related(
        'order', 'order__user', 'requested_by', 'reviewed_by'
    ).all().order_by('-requested_at')
    
    # Apply search
    if search:
        return_requests = return_requests.filter(
            Q(order__order_id__icontains=search) |
            Q(order__user__username__icontains=search) |
            Q(order__user__email__icontains=search)
        )
    
    # Apply status filter
    if status_filter:
        return_requests = return_requests.filter(status=status_filter)
    
    # Pagination
    paginator = Paginator(return_requests, 20)
    page = request.GET.get('page')
    return_requests_page = paginator.get_page(page)
    
    # Statistics
    total_requests = return_requests.count()
    pending_requests = return_requests.filter(status='pending').count()
    approved_requests = return_requests.filter(status='approved').count()
    rejected_requests = return_requests.filter(status='rejected').count()
    completed_requests = return_requests.filter(status='completed').count()
    
    context = {
        'return_requests': return_requests_page,
        'search': search,
        'status_filter': status_filter,
        'total_requests': total_requests,
        'pending_requests': pending_requests,
        'approved_requests': approved_requests,
        'rejected_requests': rejected_requests,
        'completed_requests': completed_requests,
    }
    
    return render(request, 'admin_panel/return_requests.html', context)


@superuser_required
def return_request_detail(request, pk):
    """View detailed information about a return request"""
    return_request = get_object_or_404(
        OrderReturn.objects.select_related('order', 'order__user', 'requested_by', 'reviewed_by'),
        pk=pk
    )
    
    # Get order items
    order_items = return_request.order.items.all()
    
    context = {
        'return_request': return_request,
        'order': return_request.order,
        'order_items': order_items,
    }
    
    return render(request, 'admin_panel/return_request_detail.html', context)

@superuser_required
@require_POST
def approve_return_request(request, pk):
    """✅ FIXED: Approve return request (don't refund until status changed to 'returned')"""
    return_request = get_object_or_404(OrderReturn, pk=pk)
    
    if return_request.status != 'pending':
        messages.error(request, 'This return request has already been processed.')
        return redirect('admin_panel:return_request_detail', pk=pk)
    
    admin_notes = request.POST.get('admin_notes', '').strip()
    
    try:
        with transaction.atomic():
            # ✅ FIX: Just approve, don't refund yet
            return_request.status = 'approved'
            return_request.reviewed_by = request.user
            return_request.reviewed_at = timezone.now()
            return_request.approved_at = timezone.now()
            return_request.admin_notes = admin_notes
            # Don't set refund_status yet - wait for 'returned' status
            return_request.save()
            
            # Update order status to return_approved
            order = return_request.order
            order.status = 'return_approved'
            order.save()
            
            messages.success(
                request,
                f'Return request approved! Change order status to "Returned" to complete the refund process.'
            )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        messages.error(request, f'Error processing return request: {str(e)}')
    
    return redirect('admin_panel:return_request_detail', pk=pk)


@superuser_required
@require_POST
def reject_return_request(request, pk):
    """Reject return request"""
    return_request = get_object_or_404(OrderReturn, pk=pk)
    
    if return_request.status != 'pending':
        messages.error(request, 'This return request has already been processed.')
        return redirect('admin_panel:return_request_detail', pk=pk)
    
    rejection_reason = request.POST.get('rejection_reason', '').strip()
    admin_notes = request.POST.get('admin_notes', '').strip()
    
    if not rejection_reason:
        messages.error(request, 'Please provide a rejection reason.')
        return redirect('admin_panel:return_request_detail', pk=pk)
    
    try:
        with transaction.atomic():
            # Update return request status
            return_request.status = 'rejected'
            return_request.reviewed_by = request.user
            return_request.reviewed_at = timezone.now()
            return_request.rejected_at = timezone.now()
            return_request.rejection_reason = rejection_reason
            return_request.admin_notes = admin_notes
            return_request.save()
            
            # Update order status back to delivered
            order = return_request.order
            order.status = 'delivered'
            order.save()
            
            messages.success(
                request, 
                f'Return request rejected successfully. User will be notified with the reason.'
            )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        messages.error(request, f'Error processing return request: {str(e)}')
    
    return redirect('admin_panel:return_request_detail', pk=pk)

# Add these imports at the top of your views.py
from users.models import Wallet, WalletTransaction


@superuser_required
def wallet_transactions(request):
    """Enhanced: View all wallet transactions with comprehensive filtering"""
    from users.models import WalletTransaction
    
    search = request.GET.get('search', '').strip()
    transaction_type_filter = request.GET.get('transaction_type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Base queryset
    transactions = WalletTransaction.objects.select_related(
        'wallet', 'wallet__user'
    ).all().order_by('-created_at')
    
    # Apply search - NOW INCLUDES TRANSACTION ID
    if search:
        transactions = transactions.filter(
            Q(wallet__user__username__icontains=search) |
            Q(wallet__user__email__icontains=search) |
            Q(reference_id__icontains=search) |
            Q(description__icontains=search) |
            Q(id__icontains=search)  # ✅ NEW: Search by transaction ID
        )
    
    # Apply transaction type filter
    if transaction_type_filter:
        transactions = transactions.filter(transaction_type=transaction_type_filter)
    
    # ✅ NEW: Date range filtering
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            transactions = transactions.filter(created_at__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            date_to_obj = date_to_obj.replace(hour=23, minute=59, second=59)
            transactions = transactions.filter(created_at__lte=date_to_obj)
        except ValueError:
            pass
    
    # Pagination
    paginator = Paginator(transactions, 50)
    page = request.GET.get('page')
    transactions_page = paginator.get_page(page)
    
    # Statistics
    total_credits = transactions.filter(amount__gt=0).aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0')
    
    total_debits = transactions.filter(amount__lt=0).aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0')
    
    # ✅ NEW: Transaction type breakdown
    transaction_type_counts = transactions.values('transaction_type').annotate(
        count=Count('id')
    ).order_by('-count')
    
    context = {
        'transactions': transactions_page,
        'search': search,
        'transaction_type_filter': transaction_type_filter,
        'date_from': date_from,  # ✅ NEW
        'date_to': date_to,  # ✅ NEW
        'total_credits': total_credits,
        'total_debits': abs(total_debits),
        'transaction_types': WalletTransaction.TRANSACTION_TYPES,
        'transaction_type_counts': transaction_type_counts,  # ✅ NEW
    }
    
    return render(request, 'admin_panel/wallet_transactions.html', context)


@superuser_required
def transaction_detail(request, transaction_id):
    """✅ NEW: Detailed view for a specific wallet transaction"""
    from users.models import WalletTransaction
    
    transaction = get_object_or_404(
        WalletTransaction.objects.select_related('wallet', 'wallet__user'),
        id=transaction_id
    )
    
    # Get related order if reference_id contains 'ORD'
    related_order = None
    if transaction.reference_id and 'ORD' in transaction.reference_id:
        try:
            related_order = Order.objects.get(order_id=transaction.reference_id)
        except Order.DoesNotExist:
            pass
    
    # Get user's recent transactions for context
    recent_transactions = WalletTransaction.objects.filter(
        wallet=transaction.wallet
    ).exclude(id=transaction.id).order_by('-created_at')[:5]
    
    # Get wallet balance at the time of this transaction
    previous_balance = transaction.balance_after - transaction.amount
    
    context = {
        'transaction': transaction,
        'related_order': related_order,
        'recent_transactions': recent_transactions,
        'previous_balance': previous_balance,
    }
    
    return render(request, 'admin_panel/transaction_detail.html', context)


@superuser_required
def user_wallet_detail(request, user_id):
    """Enhanced: View specific user's wallet with filters and transaction details"""
    from users.models import WalletTransaction
    
    try:
        # Get user
        user = get_object_or_404(User, id=user_id)
        
        # Get or create wallet
        wallet, created = Wallet.objects.get_or_create(user=user)
        
        # ✅ NEW: Get filter parameters
        transaction_type_filter = request.GET.get('transaction_type', '')
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        
        # Get transactions with filters
        transactions = WalletTransaction.objects.filter(
            wallet=wallet
        ).order_by('-created_at')
        
        # ✅ NEW: Apply filters
        if transaction_type_filter:
            transactions = transactions.filter(transaction_type=transaction_type_filter)
        
        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
                transactions = transactions.filter(created_at__gte=date_from_obj)
            except ValueError:
                pass
        
        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
                date_to_obj = date_to_obj.replace(hour=23, minute=59, second=59)
                transactions = transactions.filter(created_at__lte=date_to_obj)
            except ValueError:
                pass
        
        # Pagination
        paginator = Paginator(transactions, 20)
        page = request.GET.get('page')
        transactions_page = paginator.get_page(page)
        
        # Statistics
        total_credits = transactions.filter(amount__gt=0).aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0')
        
        total_debits = transactions.filter(amount__lt=0).aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0')
        
        # ✅ NEW: Transaction type breakdown
        transaction_breakdown = transactions.values('transaction_type').annotate(
            count=Count('id'),
            total_amount=Sum('amount')
        ).order_by('-count')
        
        # ✅ NEW: Count orders related to this wallet
        order_related_transactions = transactions.filter(
            reference_id__icontains='ORD'
        ).count()
        
        context = {
            'user': user,
            'wallet': wallet,
            'transactions': transactions_page,
            'total_credits': total_credits,
            'total_debits': abs(total_debits),
            'transaction_breakdown': transaction_breakdown,  # ✅ NEW
            'order_related_transactions': order_related_transactions,  # ✅ NEW
            'transaction_type_filter': transaction_type_filter,  # ✅ NEW
            'date_from': date_from,  # ✅ NEW
            'date_to': date_to,  # ✅ NEW
            'transaction_types': WalletTransaction.TRANSACTION_TYPES,  # ✅ NEW
        }
        
        return render(request, 'admin_panel/user_wallet_detail.html', context)
        
    except User.DoesNotExist:
        messages.error(request, f"User with ID {user_id} not found.")
        return redirect('admin_panel:user_list')
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        messages.error(request, f"Error loading wallet details: {str(e)}")
        return redirect('admin_panel:wallet_transactions')


@superuser_required
@require_POST
def admin_adjust_wallet(request, user_id):
    """Enhanced: Admin wallet adjustment with better tracking"""
    user = get_object_or_404(User, id=user_id)
    wallet = Wallet.objects.get_or_create(user=user)[0]
    
    action = request.POST.get('action')
    amount = request.POST.get('amount', '').strip()
    reason = request.POST.get('reason', '').strip()
    
    if not amount or not reason:
        messages.error(request, 'Please provide amount and reason.')
        return redirect('admin_panel:user_wallet_detail', user_id=user_id)
    
    try:
        amount = Decimal(amount)
        if amount <= 0:
            raise ValueError("Amount must be positive")
        
        # ✅ ENHANCED: Create unique reference ID with better tracking
        reference_id = f'ADMIN-ADJ-{timezone.now().strftime("%Y%m%d%H%M%S")}-{user_id}'
        
        if action == 'credit':
            wallet.add_money(
                amount=amount,
                transaction_type='credit_admin',
                description=f'Admin credit by {request.user.username}: {reason}',  # ✅ Track admin user
                reference_id=reference_id
            )
            messages.success(
                request, 
                f'₹{amount} credited to {user.username}\'s wallet successfully. '
                f'Reference ID: {reference_id}'  # ✅ Show reference ID
            )
        
        elif action == 'debit':
            if wallet.balance < amount:
                messages.error(
                    request, 
                    f'Insufficient wallet balance. Current balance: ₹{wallet.balance}'
                )
                return redirect('admin_panel:user_wallet_detail', user_id=user_id)
            
            wallet.deduct_money(
                amount=amount,
                transaction_type='debit_admin',
                description=f'Admin debit by {request.user.username}: {reason}',  # ✅ Track admin user
                reference_id=reference_id
            )
            messages.success(
                request, 
                f'₹{amount} debited from {user.username}\'s wallet successfully. '
                f'Reference ID: {reference_id}'  # ✅ Show reference ID
            )
        
        else:
            messages.error(request, 'Invalid action specified.')
    
    except (ValueError, Decimal.InvalidOperation):
        messages.error(request, 'Invalid amount provided. Please enter a valid number.')
    except Exception as e:
        messages.error(request, f'Error adjusting wallet: {str(e)}')
    
    return redirect('admin_panel:user_wallet_detail', user_id=user_id)


@superuser_required
def wallet_statistics(request):
    """✅ NEW: Dashboard view for wallet statistics and insights"""
    from users.models import WalletTransaction
    
    # Date range filter (default to last 30 days)
    days = int(request.GET.get('days', 30))
    start_date = timezone.now() - timedelta(days=days)
    
    # Overall statistics
    total_wallets = Wallet.objects.count()
    active_wallets = Wallet.objects.filter(balance__gt=0).count()
    total_wallet_balance = Wallet.objects.aggregate(
        total=Sum('balance')
    )['total'] or Decimal('0')
    
    # Transaction statistics for the period
    period_transactions = WalletTransaction.objects.filter(
        created_at__gte=start_date
    )
    
    total_transactions = period_transactions.count()
    total_credits = period_transactions.filter(amount__gt=0).aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0')
    total_debits = period_transactions.filter(amount__lt=0).aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0')
    
    # Transaction type breakdown
    transaction_type_breakdown = period_transactions.values(
        'transaction_type'
    ).annotate(
        count=Count('id'),
        total_amount=Sum('amount')
    ).order_by('-count')
    
    # Daily transaction trend
    daily_trend = period_transactions.annotate(
        date=TruncDate('created_at')
    ).values('date').annotate(
        credit_count=Count('id', filter=Q(amount__gt=0)),
        debit_count=Count('id', filter=Q(amount__lt=0)),
        credit_amount=Sum('amount', filter=Q(amount__gt=0)),
        debit_amount=Sum('amount', filter=Q(amount__lt=0))
    ).order_by('date')
    
    # Top users by wallet balance
    top_balance_users = Wallet.objects.select_related('user').order_by('-balance')[:10]
    
    # Top users by transaction count
    from django.db.models import Count as CountFunc
    top_active_users = Wallet.objects.annotate(
        transaction_count=CountFunc('wallettransaction')
    ).select_related('user').order_by('-transaction_count')[:10]
    
    # Recent large transactions
    large_transactions = WalletTransaction.objects.filter(
        created_at__gte=start_date
    ).select_related('wallet', 'wallet__user').order_by('-amount')[:20]
    
    context = {
        'days': days,
        'start_date': start_date,
        'total_wallets': total_wallets,
        'active_wallets': active_wallets,
        'total_wallet_balance': total_wallet_balance,
        'total_transactions': total_transactions,
        'total_credits': total_credits,
        'total_debits': abs(total_debits),
        'transaction_type_breakdown': transaction_type_breakdown,
        'daily_trend': daily_trend,
        'top_balance_users': top_balance_users,
        'top_active_users': top_active_users,
        'large_transactions': large_transactions,
    }
    
    return render(request, 'admin_panel/wallet_statistics.html', context)


# ==================== SALES REPORT VIEWS ====================

@superuser_required
def sales_report(request):
    """Main sales report view with filters and key metrics"""
    from .forms import SalesReportFilterForm
    
    # Get filter parameters
    form = SalesReportFilterForm(request.GET or None)
    
    # Default to current month if no filters
    now = timezone.now()
    start_date = None
    end_date = None
    period = request.GET.get('period', 'monthly')
    
    if form.is_valid():
        period = form.cleaned_data.get('period')
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')
    
    # Calculate date range based on period
    if period == 'daily':
        start_date = now.date()
        end_date = now.date()
    elif period == 'weekly':
        start_date = now.date() - timedelta(days=now.weekday())
        end_date = start_date + timedelta(days=6)
    elif period == 'monthly':
        start_date = now.date().replace(day=1)
        # Last day of current month
        if now.month == 12:
            end_date = now.date().replace(day=31)
        else:
            end_date = (now.replace(month=now.month + 1, day=1) - timedelta(days=1)).date()
    elif period == 'yearly':
        start_date = now.date().replace(month=1, day=1)
        end_date = now.date().replace(month=12, day=31)
    elif period == 'custom':
        # Use dates from form
        if not start_date or not end_date:
            start_date = now.date().replace(day=1)
            end_date = now.date()
    
    # Ensure dates are datetime objects for filtering
    if start_date and end_date:
        start_datetime = timezone.make_aware(datetime.combine(start_date, datetime.min.time()))
        end_datetime = timezone.make_aware(datetime.combine(end_date, datetime.max.time()))
    else:
        start_datetime = timezone.make_aware(datetime.combine(now.date().replace(day=1), datetime.min.time()))
        end_datetime = timezone.now()
    
    # Base queryset - exclude cancelled and return_requested orders
    orders = Order.objects.filter(
        created_at__gte=start_datetime,
        created_at__lte=end_datetime
    ).exclude(
        status__in=['cancelled', 'return_requested', 'returned']
    )
    
    # Calculate key metrics
    total_orders = orders.count()
    total_amount = orders.aggregate(total=Sum('total'))['total'] or Decimal('0')
    total_discount = orders.aggregate(total=Sum('discount'))['total'] or Decimal('0')
    total_shipping = orders.aggregate(total=Sum('shipping_charge'))['total'] or Decimal('0')
    
    # Calculate average order value
    avg_order_value = total_amount / total_orders if total_orders > 0 else Decimal('0')
    
    # Payment method breakdown
    cod_orders = orders.filter(payment_method='cod').count()
    online_orders = orders.filter(payment_method='razorpay').count()
    
    cod_amount = orders.filter(payment_method='cod').aggregate(total=Sum('total'))['total'] or Decimal('0')
    online_amount = orders.filter(payment_method='razorpay').aggregate(total=Sum('total'))['total'] or Decimal('0')
    
    # Status breakdown
    status_breakdown = orders.values('status').annotate(
        count=Count('id'),
        total=Sum('total')
    ).order_by('-count')
    
    # Top selling products
    top_products = OrderItem.objects.filter(
        order__created_at__gte=start_datetime,
        order__created_at__lte=end_datetime,
        order__status__in=['confirmed', 'processing', 'shipped', 'out_for_delivery', 'delivered'],
        status='active'
    ).values('product__name').annotate(
        total_quantity=Sum('quantity'),
        total_revenue=Sum('item_total')
    ).order_by('-total_quantity')[:10]
    
    # Daily sales trend (for charts)
    daily_sales = orders.annotate(
        date=TruncDate('created_at')
    ).values('date').annotate(
        total_orders=Count('id'),
        total_amount=Sum('total')
    ).order_by('date')
    
    # Coupon usage statistics
    coupon_discounts = orders.filter(coupon__isnull=False).aggregate(
        total_coupon_discount=Sum('discount'),
        coupon_usage_count=Count('id')
    )
    
    # Calculate net revenue (total - discounts)
    net_revenue = total_amount - total_discount
    
    context = {
        'form': form,
        'period': period,
        'start_date': start_date,
        'end_date': end_date,
        
        # Key metrics
        'total_orders': total_orders,
        'total_amount': total_amount,
        'total_discount': total_discount,
        'net_revenue': net_revenue,
        'avg_order_value': avg_order_value,
        'total_shipping': total_shipping,
        
        # Payment breakdown
        'cod_orders': cod_orders,
        'online_orders': online_orders,
        'cod_amount': cod_amount,
        'online_amount': online_amount,
        
        # Additional stats
        'status_breakdown': status_breakdown,
        'top_products': top_products,
        'daily_sales': daily_sales,
        'coupon_discounts': coupon_discounts,
        
        # Orders for detailed table
        'orders': orders.order_by('-created_at')[:50],  # Show latest 50
    }
    
    return render(request, 'admin_panel/sales_report.html', context)


@superuser_required
def download_sales_report_pdf(request):
    """Download sales report as PDF"""
    from .forms import SalesReportFilterForm
    
    # Get the same filters as main report
    form = SalesReportFilterForm(request.GET or None)
    
    now = timezone.now()
    period = request.GET.get('period', 'monthly')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Parse dates
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Calculate date range based on period (same logic as main view)
    if not start_date or not end_date:
        if period == 'daily':
            start_date = now.date()
            end_date = now.date()
        elif period == 'weekly':
            start_date = now.date() - timedelta(days=now.weekday())
            end_date = start_date + timedelta(days=6)
        elif period == 'monthly':
            start_date = now.date().replace(day=1)
            if now.month == 12:
                end_date = now.date().replace(day=31)
            else:
                end_date = (now.replace(month=now.month + 1, day=1) - timedelta(days=1)).date()
        elif period == 'yearly':
            start_date = now.date().replace(month=1, day=1)
            end_date = now.date().replace(month=12, day=31)
    
    start_datetime = timezone.make_aware(datetime.combine(start_date, datetime.min.time()))
    end_datetime = timezone.make_aware(datetime.combine(end_date, datetime.max.time()))
    
    # Get orders
    orders = Order.objects.filter(
        created_at__gte=start_datetime,
        created_at__lte=end_datetime
    ).exclude(
        status__in=['cancelled', 'return_requested', 'returned']
    ).order_by('-created_at')
    
    # Calculate metrics
    total_orders = orders.count()
    total_amount = orders.aggregate(total=Sum('total'))['total'] or Decimal('0')
    total_discount = orders.aggregate(total=Sum('discount'))['total'] or Decimal('0')
    net_revenue = total_amount - total_discount
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    
    # Container for PDF elements
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#667eea'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#333333'),
        spaceAfter=12,
    )
    
    # Title
    elements.append(Paragraph("WatchItUp - Sales Report", title_style))
    elements.append(Spacer(1, 12))
    
    # Report period
    period_text = f"<b>Report Period:</b> {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
    elements.append(Paragraph(period_text, styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Summary metrics table
    elements.append(Paragraph("Summary Metrics", heading_style))
    summary_data = [
        ['Metric', 'Value'],
        ['Total Orders', str(total_orders)],
        ['Total Amount', f'₹{total_amount:,.2f}'],
        ['Total Discount', f'₹{total_discount:,.2f}'],
        ['Net Revenue', f'₹{net_revenue:,.2f}'],
        ['Average Order Value', f'₹{(total_amount / total_orders if total_orders > 0 else 0):,.2f}'],
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Orders detail table
    elements.append(Paragraph("Order Details", heading_style))
    
    order_data = [['Order ID', 'Date', 'Customer', 'Amount', 'Discount', 'Status']]
    
    for order in orders[:100]:  # Limit to 100 orders for PDF
        order_data.append([
            order.order_id,
            order.created_at.strftime('%Y-%m-%d'),
            order.user.username[:15],
            f'₹{order.total:,.2f}',
            f'₹{order.discount:,.2f}',
            order.get_status_display()[:15]
        ])
    
    order_table = Table(order_data, colWidths=[1.2*inch, 1*inch, 1.2*inch, 1*inch, 1*inch, 1.2*inch])
    order_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(order_table)
    
    # Build PDF
    doc.build(elements)
    
    # Get PDF data
    pdf = buffer.getvalue()
    buffer.close()
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    filename = f'sales_report_{start_date}_{end_date}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(pdf)
    
    return response


@superuser_required
def download_sales_report_excel(request):
    """Download sales report as Excel"""
    from .forms import SalesReportFilterForm
    
    # Get the same filters as main report
    now = timezone.now()
    period = request.GET.get('period', 'monthly')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Parse dates
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Calculate date range
    if not start_date or not end_date:
        if period == 'daily':
            start_date = now.date()
            end_date = now.date()
        elif period == 'weekly':
            start_date = now.date() - timedelta(days=now.weekday())
            end_date = start_date + timedelta(days=6)
        elif period == 'monthly':
            start_date = now.date().replace(day=1)
            if now.month == 12:
                end_date = now.date().replace(day=31)
            else:
                end_date = (now.replace(month=now.month + 1, day=1) - timedelta(days=1)).date()
        elif period == 'yearly':
            start_date = now.date().replace(month=1, day=1)
            end_date = now.date().replace(month=12, day=31)
    
    start_datetime = timezone.make_aware(datetime.combine(start_date, datetime.min.time()))
    end_datetime = timezone.make_aware(datetime.combine(end_date, datetime.max.time()))
    
    # Get orders
    orders = Order.objects.filter(
        created_at__gte=start_datetime,
        created_at__lte=end_datetime
    ).exclude(
        status__in=['cancelled', 'return_requested', 'returned']
    ).order_by('-created_at')
    
    # Calculate metrics
    total_orders = orders.count()
    total_amount = orders.aggregate(total=Sum('total'))['total'] or Decimal('0')
    total_discount = orders.aggregate(total=Sum('discount'))['total'] or Decimal('0')
    net_revenue = total_amount - total_discount
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    # Styling
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = 'WatchItUp - Sales Report'
    ws['A1'].font = Font(bold=True, size=16, color="667eea")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Report period
    ws.merge_cells('A2:G2')
    ws['A2'] = f'Period: {start_date.strftime("%B %d, %Y")} to {end_date.strftime("%B %d, %Y")}'
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Summary section
    ws['A4'] = 'Summary Metrics'
    ws['A4'].font = Font(bold=True, size=14)
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Orders', total_orders],
        ['Total Amount', float(total_amount)],
        ['Total Discount', float(total_discount)],
        ['Net Revenue', float(net_revenue)],
        ['Average Order Value', float(total_amount / total_orders if total_orders > 0 else 0)],
    ]
    
    row = 5
    for data in summary_data:
        ws[f'A{row}'] = data[0]
        ws[f'B{row}'] = data[1]
        
        if row == 5:  # Header row
            ws[f'A{row}'].fill = header_fill
            ws[f'B{row}'].fill = header_fill
            ws[f'A{row}'].font = header_font
            ws[f'B{row}'].font = header_font
        
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        
        # Format currency cells
        if row > 5 and 'Amount' in data[0] or 'Revenue' in data[0]:
            ws[f'B{row}'].number_format = '₹#,##0.00'
        
        row += 1
    
    # Orders detail section
    ws[f'A{row + 2}'] = 'Order Details'
    ws[f'A{row + 2}'].font = Font(bold=True, size=14)
    
    # Headers
    headers = ['Order ID', 'Date', 'Customer', 'Email', 'Amount', 'Discount', 'Status']
    header_row = row + 3
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Order data
    data_row = header_row + 1
    for order in orders:
        ws.cell(row=data_row, column=1, value=order.order_id).border = border
        ws.cell(row=data_row, column=2, value=order.created_at.strftime('%Y-%m-%d %H:%M')).border = border
        ws.cell(row=data_row, column=3, value=order.user.username).border = border
        ws.cell(row=data_row, column=4, value=order.user.email).border = border
        
        amount_cell = ws.cell(row=data_row, column=5, value=float(order.total))
        amount_cell.number_format = '₹#,##0.00'
        amount_cell.border = border
        
        discount_cell = ws.cell(row=data_row, column=6, value=float(order.discount))
        discount_cell.number_format = '₹#,##0.00'
        discount_cell.border = border
        
        ws.cell(row=data_row, column=7, value=order.get_status_display()).border = border
        
        data_row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Create response
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'sales_report_{start_date}_{end_date}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response